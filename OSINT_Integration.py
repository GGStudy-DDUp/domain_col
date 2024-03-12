import datetime
import json
import os
import sys

from openpyxl import Workbook


# if len(sys.argv) < 2:
#     print("Usage: python xx.py <path>")
#     sys.exit(1)
#
# # 读取路径值
# PATH = sys.argv[1]

PATH = 'result/2024-02-14/'
host_collect = [['序号', '域名', '来源']]
url_collect = [['序号', 'URL', '来源']]
ips_collect = [['序号', 'IPS', '来源']]
oneforall_collect = [['序号', 'URL', '域名', 'CName', 'IP', 'CDN', '端口', '状态', '返回', '标题', 'Banner', '组织', '来源']]
host_num = 1
url_num = 1
ips_num = 1
oneforall_num = 1
# 创建一个新的Excel工作簿
wb = Workbook()
ws_oneforall = wb.active
ws_oneforall.title = "OneForAll"
ws_host = wb.create_sheet("域名")
ws_url = wb.create_sheet("URL")
ws_ips = wb.create_sheet("IPS")

# 获取当前目录下所有文件名
files = os.listdir(PATH)
# 筛选以"theHarvester_"开头且以".json"结尾的文件名
json_files = [f for f in files if f.startswith('theHarvester_') and f.endswith('.json')]

# theHarvester_*.json JSON文件整合
for json_file in json_files:
    with open(PATH + json_file, 'r') as file:
        # 加载JSON文件内容
        data = json.load(file)

    if 'hosts' in data:
        # 获取hosts键值
        for host in data['hosts']:
            host_collect.append([host_num, host.strip().replace('\n', ''), json_file])
            host_num += 1

    if 'interesting_urls' in data:
        # 获取interesting_urls键值
        for interesting_url in data['interesting_urls']:
            url_collect.append([url_num, interesting_url.strip().replace('\n', ''), json_file])
            url_num += 1

    if 'ips' in data:
        # 获取interesting_urls键值
        for ips in data['ips']:
            ips_collect.append([ips_num, ips.strip().replace('\n', ''), json_file])
            ips_num += 1

# oneforall输出整合
with open('result/2024-02-14/baidu.com.json', 'r') as file:
    # 加载JSON文件内容
    data = json.load(file)
    for i in data:
        oneforall_collect.append(
            [oneforall_num, i['url'], i['subdomain'], i['cname'], i['ip'], i['cdn'], i['port'], i['status'],
             i['reason'], i['title'], i['banner'], i['org'], i['source']])
        oneforall_num += 1

# amass_M1.txt 整合
with open(PATH + 'amass_M1.txt', 'r') as f:
    for host in f.readlines():
        if host in host_collect:
            continue
        host_collect.append([host_num, host.strip().replace('\n', ''), 'amass_M1.txt'])
        host_num += 1

# goofuzz_domain_M1.txt 整合
with open(PATH + 'goofuzz_domain_M1.txt', 'r') as f:
    for file in f.readlines():
        if file.strip() == '':
            continue
        host_collect.append([host_num, file.strip().replace('\n', ''), 'goofuzz_domain_M1.txt'])
        host_num += 1

# goofuzz_back_M1.txt 整合
with open(PATH + 'goofuzz_back_M1.txt', 'r') as f:
    for file in f.readlines():
        if file.strip() == '':
            continue
        url_collect.append([url_num, file.strip().replace('\n', ''), 'goofuzz_back_M1.txt'])
        url_num += 1

# goofuzz_file_M1.txt 整合
with open(PATH + 'goofuzz_file_M1.txt', 'r') as f:
    for file in f.readlines():
        if file.strip() == '':
            continue
        url_collect.append([url_num, file.strip().replace('\n', ''), 'goofuzz_file_M1.txt'])
        url_num += 1

# 数据写入
for tem in oneforall_collect:
    ws_oneforall.append(tem)
for tem in host_collect:
    ws_host.append(tem)
for tem in url_collect:
    ws_url.append(tem)
for tem in ips_collect:
    ws_ips.append(tem)

# 获取当前日期
current_date = datetime.datetime.now().strftime("%Y-%m-%d")
# 保存
wb.save(f"result/osint_{current_date}.xlsx")
print("数据整合完成")